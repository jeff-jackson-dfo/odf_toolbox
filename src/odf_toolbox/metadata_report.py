# Load required base libraries
import glob
import os

# Load required installed libraries
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Import required odf_toolbox libraries
from odf_toolbox.odfhdr import OdfHeader


def generate_report(file_path: str, wildcard: str, outfile: str) -> None:
    """
    Generates a report based on the metadata from ODF files as an Excel file.

    Parameters
    ----------
    file_path : str
        The file path to where the ODF files are located.
    wildcard:
        The wildcard string to filter ODF files.
    outfile:
        The output file name.

    """

    report_headings = [
        'File Name',
        'File_Spec',
        'Country_Institute_Code',
        'Cruise_Number',
        'Organization',
        'Chief_Scientist',
        'Start_Date',
        'End_Date',
        'Platform',
        'Cruise_Name',
        'Cruise_Description',
        'Data_Type',
        'Event_Number',
        'Event_Qualifier1',
        'Event_Qualifier2',
        'Creation_Date',
        'Orig_Creation_Date',
        'Start_Date_Time',
        'End_Date_Time',
        'Initial_Latitude',
        'Initial_Longitude',
        'Min_Depth',
        'Max_Depth',
        'Sampling_Interval',
        'Sounding',
        'Depth_Off_Bottom',
        'Station_Name',
        'Set_Number',
        'Event_Comments',
        'Inst_Type',
        'Model',
        'Serial_Number',
        'Description'
    ]

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Select the default sheet (usually named 'Sheet')
    worksheet = workbook.active

    # Add the report headings as the first row
    worksheet.append(report_headings)

    os.chdir(file_path)
    odfFiles = glob.glob(wildcard)

    for odf_file in odfFiles:
        odf = OdfHeader()
        odf.read_odf(file_path + odf_file)
        meta = list()
        meta.append(odf_file)
        meta.append(odf.get_file_specification().strip("'"))
        meta.append(odf.cruise_header.get_country_institute_code())
        meta.append(odf.cruise_header.get_cruise_number().strip("'"))
        meta.append(odf.cruise_header.get_organization().strip("'"))
        meta.append(odf.cruise_header.get_chief_scientist().strip("'"))
        meta.append(odf.cruise_header.get_start_date().strip("'"))
        meta.append(odf.cruise_header.get_end_date().strip("'"))
        meta.append(odf.cruise_header.get_platform().strip("'"))
        meta.append(odf.cruise_header.get_cruise_name().strip("'"))
        meta.append(odf.cruise_header.get_cruise_description().strip("'"))
        meta.append(odf.event_header.get_data_type().strip("'"))
        meta.append(odf.event_header.get_event_number().strip("'"))
        meta.append(odf.event_header.get_event_qualifier1().strip("'"))
        meta.append(odf.event_header.get_event_qualifier2().strip("'"))
        meta.append(odf.event_header.get_creation_date().strip("'"))
        meta.append(odf.event_header.get_original_creation_date().strip("'"))
        meta.append(odf.event_header.get_start_date_time().strip("'"))
        meta.append(odf.event_header.get_end_date_time().strip("'"))
        meta.append(odf.event_header.get_initial_latitude())
        meta.append(odf.event_header.get_initial_longitude())
        meta.append(odf.event_header.get_min_depth())
        meta.append(odf.event_header.get_max_depth())
        meta.append(odf.event_header.get_sampling_interval())
        meta.append(odf.event_header.get_sounding())
        meta.append(odf.event_header.get_depth_off_bottom())
        meta.append(odf.event_header.get_station_name().strip("'"))
        meta.append(odf.event_header.get_set_number().strip("'"))
        the_event_comments = ""
        for ec in odf.event_header.get_event_comments():
            the_event_comments = the_event_comments + " " + ec.strip("'")
        meta.append(the_event_comments)
        meta.append(odf.instrument_header.get_instrument_type().strip("'"))
        meta.append(odf.instrument_header.get_model().strip("'"))
        meta.append(odf.instrument_header.get_serial_number().strip("'"))
        meta.append(odf.instrument_header.get_description().strip("'"))

        # Add the metadata from the current ODF file to the report
        worksheet.append(meta)

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    max_length = max(len(str(cell.value)), max_length)
                finally:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            worksheet.column_dimensions[column].width = adjusted_width

    # Center text horizontally and vertically
    for col, cname in enumerate(worksheet.columns):
        column_letter = get_column_letter(col + 1)
        for cell in worksheet[column_letter + ":" + column_letter]:
            cell.alignment = Alignment(horizontal='center')

    # Save the workbook to a file
    workbook.save(outfile)

    # Print a success message
    print("Excel file created successfully!")


if __name__ == "__main__":
    generate_report("C:\\DEV\\CTD_Backlog\\ODF\\", "*_dn.odf", "CTD_Metadata.xlsx")